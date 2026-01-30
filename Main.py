from __future__ import annotations

import logging
import re
import time
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Sequence, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import PyPDF2
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from scipy import sparse
from scipy.integrate import simpson as simps
from scipy.sparse.linalg import spsolve
import argparse
__version__ = "0.1.0"

# ----------------------------- Configuration -----------------------------

@dataclass(frozen=True)
class Config:
    # IO
    samples_dir: Path = Path("Samples")
    results_dir: Path = Path("Results")
    file_glob: str = "*.nja"

    # Background removal (arPLS-like)
    lam_default: float = 1_000_000.0
    lam_dynamic: float = 10_000.0
    p: float = 0.001
    niter: int = 100
    dynamic_background: bool = False
    # If dynamic_background=True, use lam_dynamic for scans in the inclusive ranges below.
    dynamic_ranges: Tuple[Tuple[int, int], ...] = ((0, 0),)  # (start_sp, end_sp)

    # Duplex smoothing (keep as in original)
    smooth_weight: float = 1.0
    smooth_window_size: int = 3

    # Optional cut
    cut_enabled: bool = False
    cut_left: float = 42.0
    cut_right: float = 44.0

    # Peak search (Origin rectangle method)
    window_size_points: int = 45          # points
    local_points: int = 45
    rectangle_height_frac: float = 0.05   # fraction of max intensity after baseline removal
    peak_min_intensity: float = 5.0

    # FWHM
    steps_left_and_right: int = 35

    # Classification bounds (base peaks around known positions)
    pos_111: float = 44.34
    pos_200: float = 50.2
    bound_width: float = 0.3

    # Plotting
    save_plots: bool = True
    show_plots: bool = True   # for headless runs set False
    dpi: int = 150


CONFIG = Config()

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger("xrd-peak-optim")


# ----------------------------- Utilities -----------------------------

def get_sp_number(filename: str) -> int:
    """Extract scan order number from filenames containing _SP<number>."""
    match = re.search(r"_SP(\d+)", filename)
    if match:
        return int(match.group(1))

    nums = re.findall(r"\d+", filename)
    if nums:
        return int(nums[-1])

    return 10**18


def list_scans(samples_dir: Path, glob_pattern: str) -> List[Path]:
    files = [p for p in samples_dir.glob(glob_pattern) if p.is_file() and p.name != ".DS_Store"]
    files.sort(key=lambda p: get_sp_number(p.name))
    return files


def load_scan_xy(file_path: Path) -> Tuple[np.ndarray, np.ndarray]:
    """Load a .nja file assumed to contain at least two whitespace-separated columns: x y"""
    data = np.loadtxt(file_path)
    if data.ndim == 1:
        data = data.reshape(1, -1)
    x = data[:, 0].astype(float)
    y = data[:, 1].astype(float)
    return x, y


def parent_sample_group(sample_stem: str) -> str:
    """Use name before _SP as the sample group (batch name)."""
    return sample_stem.split("_SP")[0]


def build_output_dirs(results_dir: Path, sample_group: str) -> Tuple[Path, Path, str]:
    """Returns (single_dir, combined_dir, date_str)"""
    date_str = datetime.now().strftime("%Y-%m-%d")
    single_dir = results_dir / "Single" / date_str / sample_group
    combined_dir = results_dir / "Combined" / date_str / sample_group
    single_dir.mkdir(parents=True, exist_ok=True)
    combined_dir.mkdir(parents=True, exist_ok=True)
    return single_dir, combined_dir, date_str


# ----------------------------- Signal processing -----------------------------

def baseline_als(y_axis: np.ndarray, lam: float, p: float, niter: int = 100) -> np.ndarray:
    y_axis = np.asarray(y_axis, dtype=float)  # ensure float
    L = len(y_axis)

    # Build D as float and in CSC format (fast for solves)
    D = sparse.diags([1.0, -2.0, 1.0], [0, -1, -2], shape=(L, L - 2), format="csc", dtype=float)

    w = np.ones(L, dtype=float)

    for _ in range(niter):
        W = sparse.spdiags(w, 0, L, L, format="csc")
        Z = (W + lam * (D @ D.T)).tocsc()  # ensure CSC
        z = spsolve(Z, w * y_axis)
        w = p * (y_axis > z) + (1 - p) * (y_axis < z)

    return z


def smooth_data(data: np.ndarray, weight: float = 1.0, window_size: int = 3) -> np.ndarray:
    """Simple window mean smoothing used in the original code for duplex steels."""
    if window_size <= 1 or weight <= 0:
        return data.copy()

    smoothed = np.copy(data)
    half = window_size // 2
    for i in range(len(data)):
        start = max(0, i - half)
        end = min(len(data), i + half + 1)
        window = data[start:end]
        smoothed[i] = (1 - weight) * data[i] + weight * np.mean(window)
    return smoothed


def cut_region(x: np.ndarray, y: np.ndarray, left: float, right: float) -> Tuple[np.ndarray, np.ndarray]:
    if left > right:
        raise ValueError("cut left bound must be <= right bound")
    mask = (x >= left) & (x <= right)
    return x[mask], y[mask]


def choose_lambda(cfg: Config, sp_number: int) -> float:
    if not cfg.dynamic_background:
        return cfg.lam_default
    for start, end in cfg.dynamic_ranges:
        if start <= sp_number <= end:
            return cfg.lam_dynamic
    return cfg.lam_default


# ----------------------------- Peak detection -----------------------------

def find_candidate_peaks(x: np.ndarray, y: np.ndarray, local_points: int) -> Tuple[np.ndarray, np.ndarray]:
    """Candidate peaks: point is candidate if it's max in its local neighborhood."""
    cand_x = []
    cand_y = []
    n = len(y)
    for i in range(n):
        start = max(0, i - local_points)
        end = min(n, i + local_points + 1)
        if y[i] == np.max(y[start:end]):
            cand_x.append(x[i])
            cand_y.append(y[i])
    return np.array(cand_x), np.array(cand_y)


def filter_by_intensity(px: Sequence[float], py: Sequence[float], threshold: float) -> Tuple[List[float], List[float]]:
    fx, fy = [], []
    for x, y in zip(px, py):
        if y > threshold:
            fx.append(float(x))
            fy.append(float(y))
    return fx, fy


def rectangle_peak_test(
    x: np.ndarray,
    y: np.ndarray,
    candidate_x: Sequence[float],
    candidate_y: Sequence[float],
    width_window_size: int,
    rectangle_height: float,
) -> Tuple[List[float], List[float]]:
    true_x, true_y = [], []
    for cx, cy in zip(candidate_x, candidate_y):
        peak_idx_arr = np.where(x == cx)[0]
        if len(peak_idx_arr) == 0:
            continue
        peak_idx = int(peak_idx_arr[0])
        x1_idx = max(0, peak_idx - width_window_size // 2)
        x2_idx = min(len(x) - 1, peak_idx + width_window_size // 2)

        y1 = y[x1_idx]
        y2 = y[x2_idx]
        y_prime = max(y1, y2)
        y_local_max = float(np.max(y[x1_idx:x2_idx + 1]))
        diff = y_local_max - y_prime

        if diff > rectangle_height:
            true_x.append(float(cx))
            true_y.append(float(cy))
    return true_x, true_y


# ----------------------------- FWHM -----------------------------

def find_lowest_local_point_left(arr: np.ndarray, peak_index: int, steps_lr: int) -> Tuple[float, int]:
    left_index = peak_index
    while left_index > 0:
        increment = 10
        while left_index - increment < 0 and increment > 0:
            increment -= 1
        if increment == 0:
            break

        left_index -= increment
        if arr[left_index] == 0:
            break

        slice_start = max(0, left_index - steps_lr)
        if np.average(arr[slice_start:left_index]) < arr[left_index]:
            continue
        break
    return float(arr[left_index]), int(left_index)


def find_lowest_local_point_right(arr: np.ndarray, peak_index: int, steps_lr: int) -> Tuple[float, int]:
    right_index = peak_index
    n = len(arr)
    while right_index < n - 1:
        increment = 10
        while right_index + increment >= n and increment > 0:
            increment -= 1
        if increment == 0:
            break

        right_index += increment
        if arr[right_index] == 0:
            break

        slice_end = min(n, right_index + steps_lr)
        if np.average(arr[right_index:slice_end]) < arr[right_index]:
            continue
        break
    return float(arr[right_index]), int(right_index)


def find_fwhm(
    x: np.ndarray,
    y: np.ndarray,
    peak_value: float,
    steps_lr: int,
    left_min_idx_list: List[int],
    right_min_idx_list: List[int],
) -> Tuple[float, float, float]:
    peak_indices = np.where(y == peak_value)[0]
    if len(peak_indices) == 0:
        raise ValueError("Peak value not found in y array")
    peak_index = int(peak_indices[0])

    half_max = peak_value / 2.0

    _, left_min_idx = find_lowest_local_point_left(y, peak_index, steps_lr)
    _, right_min_idx = find_lowest_local_point_right(y, peak_index, steps_lr)
    left_min_idx_list.append(left_min_idx)
    right_min_idx_list.append(right_min_idx)

    left_idx = peak_index
    while left_idx > 0 and y[left_idx] > half_max:
        if left_idx == left_min_idx:
            break
        left_idx -= 1

    right_idx = peak_index
    while right_idx < len(y) - 1 and y[right_idx] > half_max:
        right_idx += 1

    fwhm = float(x[right_idx] - x[left_idx])
    return fwhm, float(x[left_idx]), float(x[right_idx])


# ----------------------------- Integration -----------------------------

def remove_duplicates_preserve_order(seq: Sequence[int]) -> List[int]:
    seen = set()
    out = []
    for item in seq:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def find_peak_regions(
    x: np.ndarray,
    y: np.ndarray,
    peaks_x: Sequence[float],
    left_min_idxs: Sequence[int],
    right_min_idxs: Sequence[int],
) -> List[Tuple[np.ndarray, np.ndarray]]:
    regions = []
    for i, _ in enumerate(peaks_x):
        left_bound = x[left_min_idxs[i]]
        right_bound = x[right_min_idxs[i]]
        mask = (x >= left_bound) & (x <= right_bound)
        regions.append((x[mask], y[mask]))
    return regions


def integrate_peak_areas(peak_regions: Sequence[Tuple[np.ndarray, np.ndarray]]) -> List[float]:
    areas: List[float] = []
    for x_region, y_region in peak_regions:
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=DeprecationWarning)
                area = float(simps(y_region, x_region))
            areas.append(area)
        except ValueError as e:
            log.warning("Integration failed for a region (%s). Skipping.", e)
    return areas


# ----------------------------- Peak classification -----------------------------

@dataclass
class PeakBuckets:
    base_111: List[float]
    base_200: List[float]
    expanded_111: List[float]
    expanded_200: List[float]
    extra1: List[float]
    extra2: List[float]


def classify_peaks(cfg: Config, peaks_x: Sequence[float], peaks_y: Sequence[float]) -> PeakBuckets:
    lower_111 = cfg.pos_111 - cfg.bound_width
    upper_111 = cfg.pos_111 + cfg.bound_width
    lower_200 = cfg.pos_200 - cfg.bound_width
    upper_200 = cfg.pos_200 + cfg.bound_width

    base_111: List[float] = []
    base_200: List[float] = []
    expanded_111: List[float] = []
    expanded_200: List[float] = []
    extra1: List[float] = []
    extra2: List[float] = []

    for x, y in zip(peaks_x, peaks_y):
        if lower_111 <= x <= upper_111:
            base_111.extend([x, y])
        elif lower_200 <= x <= upper_200:
            base_200.extend([x, y])
        elif x < lower_111:
            if len(expanded_111) == 0:
                expanded_111.extend([x, y])
            else:
                if len(extra1) == 0:
                    extra1.extend([x, y])
                else:
                    extra2.extend([x, y])
        elif x < lower_200:
            if len(expanded_200) == 0:
                expanded_200.extend([x, y])
            else:
                if len(extra1) == 0:
                    extra1.extend([x, y])
                else:
                    extra2.extend([x, y])

    return PeakBuckets(base_111, base_200, expanded_111, expanded_200, extra1, extra2)


def pad_bucket_to_4(bucket: List[float]) -> List[float]:
    if len(bucket) == 0:
        return [0.0, 0.0, 0.0, 0.0]
    if len(bucket) == 2:
        return [bucket[0], bucket[1], 0.0, 0.0]
    if len(bucket) >= 4:
        return bucket[:4]
    return bucket + [0.0] * (4 - len(bucket))


def update_buckets_with_metrics(
    buckets: PeakBuckets,
    true_peaks_x: Sequence[float],
    fwhm_list: Sequence[float],
    area_list: Sequence[float],
) -> PeakBuckets:
    b111 = buckets.base_111.copy()
    b200 = buckets.base_200.copy()
    e111 = buckets.expanded_111.copy()
    e200 = buckets.expanded_200.copy()
    ex1 = buckets.extra1.copy()
    ex2 = buckets.extra2.copy()

    for i, x in enumerate(true_peaks_x):
        fwhm = float(fwhm_list[i]) if i < len(fwhm_list) else 0.0
        area = float(area_list[i]) if i < len(area_list) else 0.0

        if x in b111:
            b111.extend([fwhm, area])
        elif x in b200:
            b200.extend([fwhm, area])
        elif x in e111:
            e111.extend([fwhm, area])
        elif x in e200:
            e200.extend([fwhm, area])
        elif x in ex1:
            ex1.extend([fwhm, area])
        elif x in ex2:
            ex2.extend([fwhm, area])

    return PeakBuckets(
        pad_bucket_to_4(b111),
        pad_bucket_to_4(b200),
        pad_bucket_to_4(e111),
        pad_bucket_to_4(e200),
        pad_bucket_to_4(ex1),
        pad_bucket_to_4(ex2),
    )


# ----------------------------- Plotting & Saving -----------------------------

def plot_and_save_scan(
    cfg: Config,
    x: np.ndarray,
    y_measured: np.ndarray,
    y_baseline: np.ndarray,
    y_bg_removed: np.ndarray,
    y_processed: np.ndarray,
    true_peaks_x: Sequence[float],
    true_peaks_y: Sequence[float],
    peak_regions: Sequence[Tuple[np.ndarray, np.ndarray]],
    iterations: int,
    scan_stem: str,
    single_out_dir: Path,
    pdf_merger: PyPDF2.PdfMerger,
) -> None:
    if not cfg.save_plots and not cfg.show_plots:
        return

    plt.figure()
    plt.xlabel(r"2${\Theta}$")
    plt.ylabel("Intensity")
    plt.plot(x, y_measured, label="Measured", linewidth=1)
    plt.plot(x, y_baseline, label="Estimated baseline", linestyle="--", linewidth=1, color = "black")
    # plt.plot(x, y_bg_removed, label="Measured - baseline", linewidth=1)
    plt.plot(x, y_processed, label="Smoothed (analysis)", linewidth=1, color="orange")

    if len(true_peaks_x) > 0:
        plt.scatter(true_peaks_x, true_peaks_y, color="green", label="Peaks", zorder=6)

    for i, (x_region, y_region) in enumerate(peak_regions):
        plt.fill_between(x_region, y_region, alpha=0.3, label=f"Peak {i+1}")

    plt.legend()
    plt.title(f"({iterations}) Area under each peak: {scan_stem}", fontsize=10)

    pdf_path = single_out_dir / f"{scan_stem}.pdf"
    if cfg.save_plots:
        plt.savefig(pdf_path, format="pdf", bbox_inches="tight", dpi=cfg.dpi)
        pdf_merger.append(str(pdf_path))

    if cfg.show_plots:
        plt.show()
    else:
        plt.close()


def export_results_excel(
    combined_out_dir: Path,
    sample_group: str,
    sp_numbers: Sequence[int],
    buckets_per_scan: Sequence[PeakBuckets],
) -> Path:
    rows = []
    for buckets in buckets_per_scan:
        row = (
            buckets.expanded_111
            + buckets.base_111
            + buckets.expanded_200
            + buckets.base_200
            + buckets.extra1
            + buckets.extra2
        )
        rows.append(row)

    final = np.array(rows, dtype=float)

    column_names: List[str] = []
    for _ in range(6):
        column_names.extend(["2 Theta", "Intensity", "FWHM", "Area"])

    df = pd.DataFrame(final, columns=column_names)
    df.insert(0, "Scan", list(sp_numbers))

    excel_path = combined_out_dir / f"{sample_group}.xlsx"
    df.to_excel(excel_path, index=False, startrow=1)

    wb = load_workbook(excel_path)
    sheet = wb.active
    merged_cells = [
        ("B1:E1", "Expanded Phase 111"),
        ("F1:I1", "Base Material 111"),
        ("J1:M1", "Expanded Phase 200"),
        ("N1:Q1", "Base Material 200"),
        ("R1:U1", "An extra peak?"),
        ("V1:Y1", "Extra two peaks??"),
    ]
    for merge_range, name in merged_cells:
        sheet.merge_cells(merge_range)
        cell = sheet[merge_range.split(":")[0]]
        cell.value = name
        cell.alignment = Alignment(horizontal="center")

    wb.save(excel_path)
    return excel_path


# ----------------------------- Main pipeline -----------------------------

def process_scan(
    cfg: Config,
    file_path: Path,
    iterations: int,
    single_out_dir: Path,
    pdf_merger: PyPDF2.PdfMerger,
) -> Tuple[int, PeakBuckets]:
    scan_stem = file_path.stem
    sp_number = get_sp_number(file_path.name)

    x, y = load_scan_xy(file_path)

    if float(np.mean(y)) <= 5:
        log.warning("Empty/low-intensity file skipped: %s", file_path.name)
        return sp_number, PeakBuckets([0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0])

    lam = choose_lambda(cfg, sp_number)
    baseline = baseline_als(y, lam=lam, p=cfg.p, niter=cfg.niter)
    y_bg_removed = y - baseline
    y_smooth = smooth_data(y_bg_removed, weight=cfg.smooth_weight, window_size=cfg.smooth_window_size)

    if cfg.cut_enabled:
        x, y_smooth = cut_region(x, y_smooth, cfg.cut_left, cfg.cut_right)

    rectangle_height = cfg.rectangle_height_frac * float(np.max(y_smooth))

    cand_x, cand_y = find_candidate_peaks(x, y_smooth, local_points=cfg.local_points)
    cand_x, cand_y = filter_by_intensity(cand_x, cand_y, threshold=cfg.peak_min_intensity)

    true_x, true_y = rectangle_peak_test(
        x=x,
        y=y_smooth,
        candidate_x=cand_x,
        candidate_y=cand_y,
        width_window_size=cfg.window_size_points,
        rectangle_height=rectangle_height,
    )
    true_x, true_y = filter_by_intensity(true_x, true_y, threshold=cfg.peak_min_intensity)

    left_min_idxs: List[int] = []
    right_min_idxs: List[int] = []
    fwhm_list: List[float] = []
    for pv in true_y:
        fwhm, _, _ = find_fwhm(
            x=x,
            y=y_smooth,
            peak_value=pv,
            steps_lr=cfg.steps_left_and_right,
            left_min_idx_list=left_min_idxs,
            right_min_idx_list=right_min_idxs,
        )
        fwhm_list.append(fwhm)

    left_min_idxs = remove_duplicates_preserve_order(left_min_idxs)
    right_min_idxs = remove_duplicates_preserve_order(right_min_idxs)

    k = min(len(true_x), len(true_y), len(fwhm_list), len(left_min_idxs), len(right_min_idxs))
    true_x = true_x[:k]
    true_y = true_y[:k]
    fwhm_list = fwhm_list[:k]
    left_min_idxs = left_min_idxs[:k]
    right_min_idxs = right_min_idxs[:k]

    peak_regions = find_peak_regions(x, y_smooth, true_x, left_min_idxs, right_min_idxs)
    areas = integrate_peak_areas(peak_regions)[:k]

    buckets = classify_peaks(cfg, true_x, true_y)
    buckets = update_buckets_with_metrics(buckets, true_x, fwhm_list, areas)

    plot_and_save_scan(
        cfg=cfg,
        x=x,
        y_measured=y,
        y_baseline=baseline,
        y_bg_removed=y_bg_removed,
        y_processed=y_smooth,
        true_peaks_x=true_x,
        true_peaks_y=true_y,
        peak_regions=peak_regions,
        iterations=iterations,
        scan_stem=scan_stem,
        single_out_dir=single_out_dir,
        pdf_merger=pdf_merger,
    )

    return sp_number, buckets


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Automated XRD peak analysis for in-situ diffraction data"
    )

    parser.add_argument(
        "--samples",
        type=Path,
        default=CONFIG.samples_dir,
        help="Directory containing .nja scan files (default: Samples/)",
    )

    parser.add_argument(
        "--results",
        type=Path,
        default=CONFIG.results_dir,
        help="Output directory for PDFs and Excel files (default: Results/)",
    )

    parser.add_argument(
        "--no-plots",
        action="store_true",
        help="Disable plot display and PDF generation",
    )

    parser.add_argument(
        "--cut",
        nargs=2,
        type=float,
        metavar=("LEFT", "RIGHT"),
        help="Cut the 2θ range before peak detection (e.g. --cut 42 44)",
    )

    parser.add_argument(
        "--dynamic-bg",
        action="store_true",
        help="Enable dynamic background lambda",
    )
    
    parser.add_argument(
    "--version",
    action="version",
    version=f"%(prog)s {__version__}",
    )

    return parser.parse_args()


def config_from_args(args: argparse.Namespace) -> Config:
    return Config(
        samples_dir=args.samples,
        results_dir=args.results,
        cut_enabled=args.cut is not None,
        cut_left=args.cut[0] if args.cut else CONFIG.cut_left,
        cut_right=args.cut[1] if args.cut else CONFIG.cut_right,
        dynamic_background=args.dynamic_bg,
        save_plots=not args.no_plots,
        show_plots=not args.no_plots,
    )

def main(cfg: Config = CONFIG) -> None:
    start = time.time()

    scans = list_scans(cfg.samples_dir, cfg.file_glob)
    if not scans:
        raise FileNotFoundError(f"No scan files found in {cfg.samples_dir} matching {cfg.file_glob}")

    sample_group = parent_sample_group(scans[0].stem)
    single_dir, combined_dir, _date = build_output_dirs(cfg.results_dir, sample_group)

    log.info("Found %d scan(s) in %s", len(scans), cfg.samples_dir)
    log.info("Output (single):   %s", single_dir)
    log.info("Output (combined): %s", combined_dir)

    pdf_merger = PyPDF2.PdfMerger()

    sp_numbers: List[int] = []
    buckets_all: List[PeakBuckets] = []

    for idx, scan_file in enumerate(scans, start=1):
        log.info("Processing (%d/%d): %s", idx, len(scans), scan_file.name)
        sp, buckets = process_scan(
            cfg=cfg,
            file_path=scan_file,
            iterations=idx,
            single_out_dir=single_dir,
            pdf_merger=pdf_merger,
        )
        sp_numbers.append(int(sp))
        buckets_all.append(buckets)

    combined_pdf_path = combined_dir / f"{sample_group}.pdf"
    pdf_merger.write(str(combined_pdf_path))
    pdf_merger.close()
    log.info("Wrote combined PDF: %s", combined_pdf_path)

    excel_path = export_results_excel(
        combined_out_dir=combined_dir,
        sample_group=sample_group,
        sp_numbers=sp_numbers,
        buckets_per_scan=buckets_all,
    )
    log.info("Wrote Excel: %s", excel_path)

    elapsed = time.time() - start
    log.info("Done. Processed %d scan(s) in %.2f s", len(scans), elapsed)


if __name__ == "__main__":
    args = parse_args()
    cfg = config_from_args(args)
    main(cfg)
